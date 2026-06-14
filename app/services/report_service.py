from typing import List, Optional
from datetime import datetime, timedelta
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..models import (
    DailyReport, ConstructionSite, DisposalSite, WeighingRecord,
    Violation, Settlement, TransportPermit
)
from ..utils.id_generator import generate_report_id
from ..utils.distance import haversine_distance
from ..services.notification_service_crud import send_notification
from ..models.notification import NotificationType


def generate_daily_report():
    from ..database import SessionLocal
    db = SessionLocal()
    try:
        today = datetime.utcnow().date()
        yesterday = today - timedelta(days=1)
        start = datetime.combine(yesterday, datetime.min.time())
        end = datetime.combine(yesterday, datetime.max.time())

        existing = (
            db.query(DailyReport)
            .filter(DailyReport.report_date == start)
            .first()
        )
        if existing:
            return existing

        records = (
            db.query(WeighingRecord)
            .filter(
                WeighingRecord.exit_weight_time >= start,
                WeighingRecord.exit_weight_time <= end,
            )
            .all()
        )

        total_transports = len([r for r in records if r.status == "completed"])
        total_volume = sum(r.volume or 0 for r in records if r.status == "completed")
        total_accepted_volume = sum(
            r.volume or 0 for r in records
            if r.status == "completed" and r.disposal_unload_time
        )

        total_duration = 0.0
        completed_count = 0
        for r in records:
            if r.exit_weight_time and r.disposal_unload_time:
                total_duration += (r.disposal_unload_time - r.exit_weight_time).total_seconds() / 60
                completed_count += 1
        avg_duration = total_duration / completed_count if completed_count > 0 else 0

        violations = (
            db.query(Violation)
            .filter(Violation.detected_at >= start, Violation.detected_at <= end)
            .all()
        )
        total_violations = len(violations)

        site_stats = {}
        for r in records:
            sid = r.construction_site_id
            if sid not in site_stats:
                site_stats[sid] = {"site_id": sid, "trips": 0, "volume": 0, "violations": 0}
            if r.status == "completed":
                site_stats[sid]["trips"] += 1
                site_stats[sid]["volume"] += r.volume or 0

        disposal_stats = {}
        for r in records:
            did = r.disposal_site_id
            if did not in disposal_stats:
                disposal_stats[did] = {"site_id": did, "trips": 0, "volume": 0}
            if r.status == "completed" and r.disposal_unload_time:
                disposal_stats[did]["trips"] += 1
                disposal_stats[did]["volume"] += r.volume or 0

        site_list = []
        for sid, stat in site_stats.items():
            cs = db.query(ConstructionSite).filter(ConstructionSite.id == sid).first()
            site_list.append({
                "site_id": sid,
                "site_name": cs.name if cs else "未知",
                "site_code": cs.site_code if cs else "",
                "district": cs.district if cs else "",
                "trips": stat["trips"],
                "volume": round(stat["volume"], 2),
            })

        disposal_list = []
        for did, stat in disposal_stats.items():
            ds = db.query(DisposalSite).filter(DisposalSite.id == did).first()
            disposal_list.append({
                "site_id": did,
                "site_name": ds.name if ds else "未知",
                "site_code": ds.site_code if ds else "",
                "trips": stat["trips"],
                "volume": round(stat["volume"], 2),
            })

        district_stats = {}
        for s in site_list:
            d = s["district"] or "未分区"
            if d not in district_stats:
                district_stats[d] = {"district": d, "trips": 0, "volume": 0}
            district_stats[d]["trips"] += s["trips"]
            district_stats[d]["volume"] += s["volume"]

        violation_types = {}
        for v in violations:
            vt = v.type.value
            if vt not in violation_types:
                violation_types[vt] = 0
            violation_types[vt] += 1

        report = DailyReport(
            report_id=generate_report_id(),
            report_date=start,
            total_transports=total_transports,
            total_volume=round(total_volume, 2),
            total_accepted_volume=round(total_accepted_volume, 2),
            total_violations=total_violations,
            avg_transport_duration=round(avg_duration, 2),
            site_stats=site_list,
            disposal_stats=disposal_list,
            district_stats=list(district_stats.values()),
            violation_stats=violation_types,
        )
        db.add(report)
        db.commit()
        db.refresh(report)

        return report
    finally:
        db.close()


def query_reports(
    db: Session,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    district: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
) -> List[DailyReport]:
    query = db.query(DailyReport)
    if start_date:
        query = query.filter(DailyReport.report_date >= start_date)
    if end_date:
        query = query.filter(DailyReport.report_date <= end_date)
    return query.order_by(DailyReport.report_date.desc()).offset(skip).limit(limit).all()


def aggregate_stats(
    db: Session,
    start_date: datetime,
    end_date: datetime,
    district: Optional[str] = None,
) -> dict:
    reports = query_reports(db, start_date, end_date)

    total_transports = sum(r.total_transports for r in reports)
    total_volume = sum(r.total_volume for r in reports)
    total_accepted = sum(r.total_accepted_volume for r in reports)
    total_violations = sum(r.total_violations for r in reports)
    avg_duration = (
        sum(r.avg_transport_duration for r in reports) / len(reports)
        if reports else 0
    )

    agg_site = {}
    agg_disposal = {}
    agg_district = {}
    agg_violations = {}

    for r in reports:
        for s in r.site_stats or []:
            if district and s.get("district") != district:
                continue
            sid = s["site_id"]
            if sid not in agg_site:
                agg_site[sid] = {
                    "site_id": sid, "site_name": s["site_name"],
                    "site_code": s["site_code"], "district": s.get("district", ""),
                    "trips": 0, "volume": 0,
                }
            agg_site[sid]["trips"] += s["trips"]
            agg_site[sid]["volume"] += s["volume"]

        for d in r.disposal_stats or []:
            did = d["site_id"]
            if did not in agg_disposal:
                agg_disposal[did] = {
                    "site_id": did, "site_name": d["site_name"],
                    "site_code": d["site_code"], "trips": 0, "volume": 0,
                }
            agg_disposal[did]["trips"] += d["trips"]
            agg_disposal[did]["volume"] += d["volume"]

        for ds in r.district_stats or []:
            if district and ds["district"] != district:
                continue
            dn = ds["district"]
            if dn not in agg_district:
                agg_district[dn] = {"district": dn, "trips": 0, "volume": 0}
            agg_district[dn]["trips"] += ds["trips"]
            agg_district[dn]["volume"] += ds["volume"]

        for vt, cnt in (r.violation_stats or {}).items():
            if vt not in agg_violations:
                agg_violations[vt] = 0
            agg_violations[vt] += cnt

    return {
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        },
        "summary": {
            "total_transports": total_transports,
            "total_volume": round(total_volume, 2),
            "total_accepted_volume": round(total_accepted, 2),
            "total_violations": total_violations,
            "avg_transport_duration_minutes": round(avg_duration, 2),
        },
        "by_site": list(agg_site.values()),
        "by_disposal": list(agg_disposal.values()),
        "by_district": list(agg_district.values()),
        "by_violation_type": agg_violations,
    }


def export_stats_to_excel(stats: dict) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "汇总统计"
    summary = stats["summary"]
    ws1.append(["指标", "数值"])
    headers = ["总运输车次", "总出土量(立方米)", "总消纳量(立方米)", "总违规次数", "平均运输周期(分钟)"]
    values = [
        summary["total_transports"],
        summary["total_volume"],
        summary["total_accepted_volume"],
        summary["total_violations"],
        summary["avg_transport_duration_minutes"],
    ]
    for h, v in zip(headers, values):
        ws1.append([h, v])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws2 = wb.create_sheet("工地统计")
    ws2.append(["工地编号", "工地名称", "区域", "车次", "出土量(立方米)"])
    for s in stats["by_site"]:
        ws2.append([s["site_code"], s["site_name"], s.get("district", ""), s["trips"], round(s["volume"], 2)])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws3 = wb.create_sheet("消纳场统计")
    ws3.append(["消纳场编号", "消纳场名称", "接收车次", "接收量(立方米)"])
    for d in stats["by_disposal"]:
        ws3.append([d["site_code"], d["site_name"], d["trips"], round(d["volume"], 2)])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws4 = wb.create_sheet("区域统计")
    ws4.append(["区域", "车次", "出土量(立方米)"])
    for ds in stats["by_district"]:
        ws4.append([ds["district"], ds["trips"], round(ds["volume"], 2)])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    ws5 = wb.create_sheet("违规统计")
    ws5.append(["违规类型", "次数"])
    for vt, cnt in stats["by_violation_type"].items():
        type_map = {
            "off_route": "偏离路线", "overtime": "超时未到达",
            "overload": "超载", "unsealed": "未密闭运输",
            "no_permit": "无准运证", "expired_permit": "准运证过期",
        }
        ws5.append([type_map.get(vt, vt), cnt])
    for cell in ws5[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
